#!/usr/bin/env python3
"""Generate a closed-bug test matrix for MetaMask Mobile (last 30 days)."""

from __future__ import annotations

import datetime as dt
import json
import re
import subprocess
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = "MetaMask/metamask-mobile"
LOOKBACK_DAYS = 30
MAX_ISSUES = 1000

OUTPUT_DIR = Path("docs/bug-test-matrix")
MD_PATH = OUTPUT_DIR / "metamask-mobile-bugs-last-30-days.md"
XLSX_PATH = OUTPUT_DIR / "metamask-mobile-bugs-last-30-days.xlsx"


def run_gh_issue_list() -> list[dict]:
    cmd = [
        "gh",
        "issue",
        "list",
        "--repo",
        REPO,
        "--state",
        "closed",
        "--limit",
        str(MAX_ISSUES),
        "--json",
        (
            "number,title,createdAt,updatedAt,closedAt,state,stateReason,labels,url,body,"
            "comments,closedByPullRequestsReferences"
        ),
    ]
    out = subprocess.check_output(cmd, text=True)
    return json.loads(out)


def clean_title(title: str) -> str:
    t = title.strip()
    t = re.sub(r"^\[[^\]]+\]\s*:?\s*", "", t)
    t = re.sub(r"^(bug|ui)\s*:\s*", "", t, flags=re.IGNORECASE)
    return t


def clean_markdown_text(text: str) -> str:
    if not text:
        return ""
    value = text
    value = re.sub(r"```.*?```", " ", value, flags=re.DOTALL)
    value = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", value)
    value = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", value)
    value = re.sub(r"<img[^>]*>", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"</?[^>]+>", " ", value)
    value = re.sub(r"`([^`]+)`", r"\1", value)
    value = value.replace("**", "").replace("__", "").replace("*", "").replace("_", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_header(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[_*`:#-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current = "__preamble__"
    sections[current] = []

    for raw_line in body.splitlines():
        stripped = raw_line.strip()
        heading_match = re.match(r"^\s*#{1,6}\s+(.+?)\s*$", stripped)
        bold_heading_match = re.match(r"^\s*\*\*(.+?)\*\*\s*$", stripped)
        if heading_match:
            current = normalize_header(heading_match.group(1))
            sections.setdefault(current, [])
            continue
        if bold_heading_match:
            current = normalize_header(bold_heading_match.group(1))
            sections.setdefault(current, [])
            continue
        sections.setdefault(current, []).append(stripped)
    return {k: "\n".join(v).strip() for k, v in sections.items()}


def first_non_empty(values: Iterable[str]) -> str:
    for value in values:
        if value and value.strip():
            return value.strip()
    return ""


def section_text(sections: dict[str, str], aliases: list[str]) -> str:
    aliases_norm = {normalize_header(a) for a in aliases}
    for key, value in sections.items():
        if normalize_header(key) in aliases_norm and value:
            return value
    return ""


def shorten(text: str, max_chars: int) -> str:
    clean = clean_markdown_text(text)
    if len(clean) <= max_chars:
        return clean
    return clean[: max_chars - 1].rstrip() + "…"


def extract_component_paths(body: str) -> list[str]:
    matches = re.findall(r"`([^`]+(?:/[A-Za-z0-9_.-]+)+)`", body)
    matches.extend(
        re.findall(
            r"\b(?:app|tests|ios|android|wdio)/[A-Za-z0-9_./-]+",
            body,
        )
    )
    seen: set[str] = set()
    ordered: list[str] = []
    for item in matches:
        text = item.strip()
        if "/" in text and len(text) < 180 and text not in seen:
            seen.add(text)
            ordered.append(text)
    return ordered


def severity_rank(severity_labels: list[str]) -> tuple[int, str]:
    if not severity_labels:
        return (99, "Unlabeled")
    candidates = []
    for label in severity_labels:
        low = label.lower()
        if "sev0" in low:
            candidates.append((0, label))
        elif "sev1" in low:
            candidates.append((1, label))
        elif "sev2" in low or "sev-2" in low:
            candidates.append((2, label))
        elif "sev3" in low:
            candidates.append((3, label))
        else:
            candidates.append((50, label))
    candidates.sort(key=lambda x: x[0])
    return candidates[0]


def infer_component_hint(text: str) -> str:
    lowered = text.lower()
    rules = [
        (r"perps|predict", "Perps/Predict view components"),
        (r"nft|erc-721|collectible", "NFT detail/send confirmation views"),
        (r"token detail|asset detail|portfolio|balance|token", "Asset list/detail components"),
        (r"explore|trending|sites|browser", "Explore/Sites view components"),
        (r"network icon|network switcher|chain", "Network switcher / chain badge components"),
        (r"copy address|clipboard|paste", "Input/copy interaction components"),
        (r"swap|bridge|quote|gas", "Swap/Bridge quote and review views"),
        (r"onboard|onboarding|import wallet|srp|recovery phrase", "Onboarding flow views"),
        (r"hardware wallet|ledger|trezor|qr scanner|scan", "Hardware wallet scanner/connect views"),
        (r"transaction|confirm|approval|signature", "Transaction confirmation views"),
        (r"notification|push", "Notification center/badge components"),
    ]
    for pattern, hint in rules:
        if re.search(pattern, lowered):
            return hint
    return "Relevant screen/view component for the reported flow"


def infer_assertion(text: str) -> str:
    lowered = text.lower()
    if re.search(r"disappear|missing|invisible|no .*icon|not shown|doesn.?t show", lowered):
        return "Assert expected elements stay visible and all intended options/icons remain rendered."
    if re.search(r"wrong|incorrect|mismatch|different|other token|another token", lowered):
        return "Assert displayed identifiers/amounts/labels exactly match the selected source state."
    if re.search(r"unable to|cannot|can.?t|fails to|failing", lowered):
        return "Assert the user action completes successfully and the UI updates to the expected next state."
    if re.search(r"paste|clipboard", lowered):
        return "Assert pasted clipboard text is accepted and bound to the input/query state."
    if re.search(r"freeze|crash|stuck|hang", lowered):
        return "Assert no freeze/crash path and that controls remain interactive after the action."
    if re.search(r"timeout|transaction not found|error", lowered):
        return "Assert error state is explicit and retry/recovery controls are available."
    if re.search(r"spacing|padding|alignment|text differs|copy text", lowered):
        return "Assert UI copy/style tokens match spec and remain consistent across equivalent views."
    return "Assert the rendered state and behavior match expected outcome for the reported scenario."


def infer_mocks(text: str, component_hint: str) -> str:
    lowered = text.lower()
    base = f"Render {component_hint} with controller/store state fixtures."
    if re.search(r"swap|bridge|quote|gas|price|rate|fee", lowered):
        return base + " Mock quote/fee API responses (success + edge case) and verify derived totals."
    if re.search(r"nft|erc-721|token id|collectible", lowered):
        return base + " Mock NFT/token metadata and deterministic IDs for selection vs confirmation."
    if re.search(r"explore|search|paste|url|sites", lowered):
        return base + " Mock search providers and clipboard input events."
    if re.search(r"network|chain|icon", lowered):
        return base + " Mock multichain network metadata and icon mapping selectors."
    if re.search(r"hardware wallet|qr scanner|scan|camera", lowered):
        return base + " Mock scanner module output and permission state."
    if re.search(r"transaction|confirm|approval|signature", lowered):
        return base + " Mock transaction controller state transitions and error branches."
    return base + " Mock any async selector/API used by the flow."


def infer_needs_e2e(text: str) -> bool:
    lowered = text.lower()
    return bool(
        re.search(
            r"hardware wallet|qr scanner|camera|deeplink|deep link|walletconnect|"
            r"push notification|biometric|background|app launch|nfc|bluetooth|permissions",
            lowered,
        )
    )


def extract_steps_snippet(steps_text: str, title_text: str) -> str:
    if not steps_text:
        return f"Trigger the user action described in '{title_text}'."
    step_source = re.sub(r"(\d+\.)", r"\n\1", steps_text)
    raw_lines = [clean_markdown_text(line) for line in step_source.splitlines()]
    lines = []
    for line in raw_lines:
        line = re.sub(r"^\d+[\).\s-]*", "", line).strip("- ").strip()
        if line and line not in {"_No response_", "No response", "TBD"}:
            lines.append(line)
    if not lines:
        return f"Trigger the user action described in '{title_text}'."
    return " -> ".join(lines[:3])


def build_problem_summary(title: str, sections: dict[str, str]) -> str:
    describe = section_text(
        sections,
        ["Describe the bug", "Bug Description", "Description", "Current Behavior"],
    )
    expected = section_text(sections, ["Expected behavior", "Expected Behavior"])
    bits = [clean_title(title)]
    if describe:
        bits.append(shorten(describe, 220))
    elif expected:
        bits.append(f"Expected: {shorten(expected, 180)}")
    return shorten(" — ".join(bits), 340)


def markdown_escape(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def normalize_title_key(title: str) -> str:
    text = clean_title(title).lower()
    text = re.sub(r"\(android.*?\)|\(ios.*?\)", " ", text)
    text = re.sub(r"v\d+\.\d+(?:\.\d+)?", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_auto_rca_comment(comment_body: str, author_login: str) -> bool:
    return (
        author_login == "github-actions"
        and "This issue has been closed. Please complete this RCA form" in comment_body
    )


def manual_comments_text(issue: dict) -> str:
    chunks: list[str] = []
    for comment in issue.get("comments") or []:
        body = comment.get("body") or ""
        author = (comment.get("author") or {}).get("login") or ""
        if is_auto_rca_comment(body, author):
            continue
        chunks.append(body)
    return "\n".join(chunks).strip()


def classify_realness(issue: dict, labels: list[str], combined_text: str, title_cluster: dict) -> tuple[str, str]:
    low_text = combined_text.lower()
    low_labels = [l.lower() for l in labels]
    state_reason = (issue.get("stateReason") or "").upper()

    if re.search(
        r"\btesting customfield mapping\b|\btesting release field mapping\b|^testing\b|\btest issue\b",
        low_text,
    ):
        return (
            "Not a real bug (test/admin issue)",
            "Issue content indicates testing/admin validation rather than product defect.",
        )

    if re.search(
        r"not a bug|cannot reproduce|not reproducible|"
        r"as designed|by design|won[’']?t fix|invalid issue",
        low_text,
    ):
        return (
            "Not a real bug (closed as expected/non-repro)",
            "Issue text/comments contain explicit non-bug signal (expected behavior/non-repro).",
        )

    duplicate_signal = bool(
        re.search(r"\bduplicate\b|\bdupe\b|same as #\d+|closing in favor of #\d+", low_text)
    ) or ("duplicate" in low_labels)

    if state_reason == "COMPLETED":
        return ("Confirmed real bug", "State reason is COMPLETED.")

    completed_siblings = [
        n for n in (title_cluster.get("completed_numbers") or []) if n != issue.get("number")
    ]
    if state_reason == "NOT_PLANNED" and completed_siblings:
        return (
            "Likely real bug (duplicate/tracked by sibling issue)",
            f"Same normalized title has completed issue(s): {', '.join(f'#{n}' for n in completed_siblings[:3])}.",
        )

    if state_reason == "NOT_PLANNED" and duplicate_signal:
        return (
            "Likely real bug (duplicate/tracked elsewhere)",
            "NOT_PLANNED plus duplicate-style signal in content.",
        )

    if state_reason == "NOT_PLANNED" and re.search(r"\bjira\b|story|ticket|tracked in|moved to", low_text):
        return (
            "Likely real bug (tracked externally)",
            "NOT_PLANNED with Jira/ticket tracking signal.",
        )

    if state_reason == "NOT_PLANNED" and any(
        l.startswith("sev") or l.startswith("regression") or l == "release-blocker" for l in low_labels
    ):
        return (
            "Likely real bug (closed without direct fix reference)",
            "NOT_PLANNED but severity/regression labels suggest valid defect report.",
        )

    return ("Unclear from issue content", "Insufficient explicit evidence in issue body/comments.")


def classify_resolution_layer(
    issue: dict,
    labels: list[str],
    combined_text: str,
    realness: str,
    title_cluster: dict,
) -> tuple[str, str]:
    if realness.startswith("Not a real bug"):
        return ("N/A", "Issue classified as non-bug.")

    closing_prs = issue.get("closedByPullRequestsReferences") or []
    if closing_prs:
        numbers = [str(pr.get("number")) for pr in closing_prs if pr.get("number")]
        return (
            "Mobile codebase",
            f"Closed by PR reference(s) in {REPO}: {', '.join('#' + n for n in numbers[:3])}.",
        )

    low_text = combined_text.lower()
    api_signals = bool(
        re.search(
            r"\bapi\b|backend|server[- ]side|indexer|quote service|market data|coingecko|relay|endpoint|"
            r"portfolio data|price feed|provider response|rpc|cap\s*&\s*vol|market cap|"
            r"volume data|trending tokens.*limited|symbol for trending token|price data",
            low_text,
        )
    )
    mobile_signals = bool(
        re.search(
            r"screen|view|modal|button|navigation|bottom sheet|app/components|"
            r"search bar|token detail|activity page|network switcher",
            low_text,
        )
    )

    completed_siblings = [
        n for n in (title_cluster.get("completed_numbers") or []) if n != issue.get("number")
    ]
    completed_sibling_with_pr = [
        n for n in (title_cluster.get("completed_with_pr") or []) if n != issue.get("number")
    ]
    if completed_sibling_with_pr:
        return (
            "Mobile codebase (via sibling issue)",
            f"Similar issue was completed with mobile PR-linked ticket(s): {', '.join(f'#{n}' for n in completed_sibling_with_pr[:3])}.",
        )
    if completed_siblings:
        return (
            "Likely mobile (via sibling issue)",
            f"Similar issue has completed sibling ticket(s): {', '.join(f'#{n}' for n in completed_siblings[:3])}.",
        )

    if api_signals and mobile_signals:
        return ("Mixed/unclear (mobile + API signals)", "Issue text contains both API and mobile UI signals.")
    if api_signals:
        return ("Likely API/backend", "Issue text/comments contain API/backend/data-provider signals.")
    if mobile_signals:
        return ("Likely mobile", "Issue text focuses on in-app UI/navigation behavior.")
    return ("Unknown", "No direct PR linkage and no strong mobile/API signal.")


def extract_mobile_impact(combined_text: str, resolution_layer: str) -> str:
    low_text = combined_text.lower()
    mobile_signal = bool(
        re.search(
            r"\bios\b|\bandroid\b|mobile app|screen|view|button|navigation|network switcher|token detail|app/",
            low_text,
        )
    )
    if resolution_layer == "N/A":
        return "No (non-bug)"
    if resolution_layer.startswith("Likely API") and not mobile_signal:
        return "Indirect/unclear"
    if mobile_signal:
        return "Yes"
    if resolution_layer.startswith("Mobile"):
        return "Yes"
    return "Unclear"


def derive_ai_validity(
    issue: dict,
    realness: str,
    realness_evidence: str,
    resolution_layer: str,
    title_cluster: dict,
) -> tuple[str, str, str]:
    issue_number = issue.get("number")
    state_reason = (issue.get("stateReason") or "").upper()
    prs = issue.get("closedByPullRequestsReferences") or []
    sibling_count = max(0, len(title_cluster.get("numbers") or []) - 1)

    if realness.startswith("Not a real bug"):
        return (
            "Invalid (non-product bug)",
            "High",
            f"{realness_evidence} State reason: {state_reason}.",
        )

    if realness == "Confirmed real bug":
        if prs:
            pr_nums = ", ".join(f"#{pr.get('number')}" for pr in prs[:3] if pr.get("number"))
            return (
                "Valid bug",
                "High",
                f"Closed as COMPLETED and linked to mobile PR(s) {pr_nums}.",
            )
        return (
            "Valid bug",
            "Medium-High",
            "Closed as COMPLETED without explicit linked PR in issue metadata.",
        )

    if realness.startswith("Likely real bug"):
        confidence = "Medium"
        if "sibling issue" in realness.lower() and sibling_count > 0:
            confidence = "Medium-High"
        if "tracked externally" in realness.lower():
            confidence = "Medium"
        return (
            "Likely valid bug",
            confidence,
            f"{realness_evidence} Cluster sibling count: {sibling_count}.",
        )

    return (
        "Needs manual review",
        "Low",
        f"Ambiguous closure context. Realness signal: {realness}. Resolution layer: {resolution_layer}.",
    )


def choose_canonical_issue(title_cluster: dict, fallback_issue_number: int) -> int:
    candidates = (
        title_cluster.get("completed_with_pr")
        or title_cluster.get("completed_numbers")
        or title_cluster.get("numbers")
        or [fallback_issue_number]
    )
    return sorted(candidates)[0]


def recommend_test_implementation(
    issue_number: int,
    ai_validity_verdict: str,
    resolution_layer: str,
    component_hint: str,
    cvt_mocks: str,
    cvt_steps: str,
    cvt_assertions: str,
    needs_e2e: bool,
    canonical_issue: int,
    title_cluster: dict,
) -> tuple[str, str]:
    sibling_count = max(0, len(title_cluster.get("numbers") or []) - 1)

    if ai_validity_verdict.startswith("Invalid"):
        return (
            "No new product regression test",
            "Keep a triage guardrail only (classification/unit check for bug-template hygiene).",
        )

    duplicate_scope = canonical_issue != issue_number and sibling_count > 0

    if resolution_layer.startswith("Likely API/backend"):
        return (
            "API integration/contract test + mobile CVT fallback",
            "API: validate provider payload/ordering/values. "
            "Mobile CVT: render error/loading/edge-state in component and assert graceful UX.",
        )

    if resolution_layer.startswith("Mixed/unclear"):
        return (
            "CVT + integration test + one E2E smoke",
            "CVT for UI regression, integration test around controller/network boundary, "
            "and one E2E to validate the end-to-end dependency chain.",
        )

    if needs_e2e:
        base = (
            f"CVT-first in {component_hint}: Given {cvt_mocks} When {cvt_steps} Then {cvt_assertions}"
        )
        extra = " Add one focused E2E for device/integration boundary."
        if duplicate_scope:
            extra += f" Reuse canonical scenario from issue #{canonical_issue} with parametrized fixtures."
        return ("Component View Test + targeted E2E", shorten(base + extra, 420))

    plan = f"CVT in {component_hint}: Given {cvt_mocks} When {cvt_steps} Then {cvt_assertions}"
    if duplicate_scope:
        plan += f" Reuse canonical test from issue #{canonical_issue} and add this case as a parameterized fixture."
    return ("Component View Test (primary)", shorten(plan, 420))


def build_title_clusters(issues: list[dict]) -> dict[str, dict]:
    clusters: dict[str, dict] = defaultdict(
        lambda: {"numbers": [], "completed_numbers": [], "completed_with_pr": []}
    )
    for issue in issues:
        key = normalize_title_key(issue["title"])
        if not key:
            continue
        cluster = clusters[key]
        cluster["numbers"].append(issue["number"])
        if (issue.get("stateReason") or "").upper() == "COMPLETED":
            cluster["completed_numbers"].append(issue["number"])
            if issue.get("closedByPullRequestsReferences"):
                cluster["completed_with_pr"].append(issue["number"])
    return clusters


def build_rows(raw_issues: list[dict]) -> list[dict]:
    now = dt.datetime.now(dt.timezone.utc)
    cutoff = now - dt.timedelta(days=LOOKBACK_DAYS)

    scoped_issues = []
    for issue in raw_issues:
        labels = [label["name"] for label in issue.get("labels") or []]
        if issue.get("state") != "CLOSED":
            continue
        if "type-bug" not in labels:
            continue
        created_at = dt.datetime.fromisoformat(issue["createdAt"].replace("Z", "+00:00"))
        if created_at < cutoff:
            continue
        scoped_issues.append(issue)

    title_clusters = build_title_clusters(scoped_issues)

    rows: list[dict] = []
    for issue in scoped_issues:
        labels = [label["name"] for label in issue.get("labels") or []]
        teams = [name for name in labels if name.lower().startswith("team-")]
        sevs = [name for name in labels if name.lower().startswith("sev")]
        sev_priority, sev_top = severity_rank(sevs)

        body = issue.get("body") or ""
        comments_text = manual_comments_text(issue)
        sections = parse_sections(body)
        summary = build_problem_summary(issue["title"], sections)

        affected = section_text(
            sections,
            ["Affected Component", "Affected Components", "Affected component(s)"],
        )
        component_paths = extract_component_paths(body)
        component_hint = first_non_empty(
            [
                ", ".join(component_paths[:2]),
                shorten(affected, 150),
                infer_component_hint(issue["title"] + " " + body),
            ]
        )

        steps = section_text(sections, ["Steps to reproduce", "Steps"])
        expected = section_text(sections, ["Expected behavior", "Expected Behavior"])
        describe = section_text(
            sections,
            ["Describe the bug", "Bug Description", "Description", "Current Behavior"],
        )
        steps_snippet = shorten(extract_steps_snippet(steps, issue["title"]), 260)
        focus_text = " ".join([issue["title"], describe, expected, steps])
        assertion = infer_assertion(focus_text)
        mocks = infer_mocks(" ".join([issue["title"], describe, expected]), component_hint)
        needs_e2e = infer_needs_e2e(focus_text)

        primary_test = (
            "Component View Test + E2E smoke (integration-sensitive)"
            if needs_e2e
            else "Component View Test"
        )
        e2e_fallback = (
            "Cover full device flow with real integration boundary (permissions/scanner/deeplink/network), and verify final user-visible state."
            if needs_e2e
            else "Optional: run one E2E regression on the same path to confirm controller + navigation wiring."
        )

        title_key = normalize_title_key(issue["title"])
        title_cluster = title_clusters.get(title_key, {})
        canonical_issue = choose_canonical_issue(title_cluster, issue["number"])
        combined_for_review = " ".join([issue["title"], body, comments_text])
        realness, realness_evidence = classify_realness(issue, labels, combined_for_review, title_cluster)
        resolution_layer, resolution_evidence = classify_resolution_layer(
            issue,
            labels,
            combined_for_review,
            realness,
            title_cluster,
        )
        mobile_impact = extract_mobile_impact(combined_for_review, resolution_layer)
        ai_validity_verdict, ai_validity_confidence, ai_validity_analysis = derive_ai_validity(
            issue,
            realness,
            realness_evidence,
            resolution_layer,
            title_cluster,
        )
        recommended_test_type, recommended_test_implementation = recommend_test_implementation(
            issue_number=issue["number"],
            ai_validity_verdict=ai_validity_verdict,
            resolution_layer=resolution_layer,
            component_hint=component_hint,
            cvt_mocks=shorten(mocks, 220),
            cvt_steps=steps_snippet,
            cvt_assertions=shorten(assertion, 180),
            needs_e2e=needs_e2e,
            canonical_issue=canonical_issue,
            title_cluster=title_cluster,
        )

        metadata_gaps = []
        if not teams:
            metadata_gaps.append("Missing team label")
        if not sevs:
            metadata_gaps.append("Missing severity label")

        row = {
            "issue_number": issue["number"],
            "issue_url": issue["url"],
            "created_at": issue["createdAt"],
            "closed_at": issue.get("closedAt") or "",
            "state_reason": issue.get("stateReason") or "",
            "team_labels": ", ".join(teams) if teams else "Unlabeled",
            "severity_labels": ", ".join(sevs) if sevs else "Unlabeled",
            "severity_top": sev_top,
            "severity_priority": sev_priority,
            "problem_summary": summary,
            "real_bug_assessment": realness,
            "real_bug_evidence": shorten(realness_evidence, 220),
            "ai_validity_verdict": ai_validity_verdict,
            "ai_validity_confidence": ai_validity_confidence,
            "ai_validity_analysis": shorten(ai_validity_analysis, 280),
            "mobile_impact": mobile_impact,
            "resolution_layer": resolution_layer,
            "resolution_evidence": shorten(resolution_evidence, 220),
            "canonical_issue": canonical_issue,
            "component_candidate": component_hint,
            "cvt_preconditions_mocks": shorten(mocks, 320),
            "cvt_steps": steps_snippet,
            "cvt_assertions": shorten(assertion, 220),
            "primary_test_type": primary_test,
            "e2e_fallback": shorten(e2e_fallback, 240),
            "recommended_test_type": recommended_test_type,
            "recommended_test_implementation": recommended_test_implementation,
            "metadata_gaps": ", ".join(metadata_gaps) if metadata_gaps else "",
        }
        rows.append(row)

    rows.sort(
        key=lambda item: (item["severity_priority"], item["created_at"]),
        reverse=False,
    )
    grouped: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[row["severity_priority"]].append(row)
    ordered_rows: list[dict] = []
    for priority in sorted(grouped.keys()):
        ordered_rows.extend(sorted(grouped[priority], key=lambda x: x["created_at"], reverse=True))
    return ordered_rows


def write_markdown(rows: list[dict], generated_at: dt.datetime) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    sev_counter = Counter(row["severity_top"] for row in rows)
    realness_counter = Counter(row["real_bug_assessment"] for row in rows)
    ai_verdict_counter = Counter(row["ai_validity_verdict"] for row in rows)
    resolution_counter = Counter(row["resolution_layer"] for row in rows)
    test_type_counter = Counter(row["recommended_test_type"] for row in rows)
    team_counter = Counter()
    for row in rows:
        for team in [t.strip() for t in row["team_labels"].split(",") if t.strip() and t.strip() != "Unlabeled"]:
            team_counter[team] += 1

    header = [
        "# MetaMask Mobile closed bug-to-test matrix (last 30 days)",
        "",
        f"- Generated at (UTC): {generated_at.isoformat(timespec='seconds')}",
        f"- Scope: `{REPO}` issues",
        f"- Filters: `state:CLOSED`, `label:type-bug`, `createdAt >= now-30d`",
        f"- Total bugs in scope: **{len(rows)}**",
        "",
        "## Severity distribution",
        "",
        "| Severity | Count |",
        "|---|---:|",
    ]
    for severity, count in sev_counter.most_common():
        header.append(f"| {severity} | {count} |")

    header.extend(
        [
            "",
            "## Real bug assessment (from issue content/comments)",
            "",
            "| Assessment | Count |",
            "|---|---:|",
        ]
    )
    for assessment, count in realness_counter.most_common():
        header.append(f"| {assessment} | {count} |")

    header.extend(
        [
            "",
            "## AI validity verdict (final per issue)",
            "",
            "| Verdict | Count |",
            "|---|---:|",
        ]
    )
    for verdict, count in ai_verdict_counter.most_common():
        header.append(f"| {verdict} | {count} |")

    header.extend(
        [
            "",
            "## Resolution layer assessment",
            "",
            "| Resolution layer | Count |",
            "|---|---:|",
        ]
    )
    for layer, count in resolution_counter.most_common():
        header.append(f"| {layer} | {count} |")

    header.extend(
        [
            "",
            "## Recommended test implementation distribution",
            "",
            "| Test implementation | Count |",
            "|---|---:|",
        ]
    )
    for test_impl, count in test_type_counter.most_common():
        header.append(f"| {test_impl} | {count} |")

    header.extend(
        [
            "",
            "## Top team labels in scoped bugs",
            "",
            "| Team label | Count |",
            "|---|---:|",
        ]
    )
    for team, count in team_counter.most_common(15):
        header.append(f"| {team} | {count} |")

    table_headers = [
        "Issue",
        "Created (UTC)",
        "Closed (UTC)",
        "State reason",
        "Team labels",
        "Severity labels",
        "Real bug assessment",
        "Real bug evidence",
        "AI validity verdict",
        "AI confidence",
        "AI analysis",
        "Applies to mobile",
        "Resolution layer",
        "Resolution evidence",
        "Canonical issue for test",
        "Recommended test implementation",
        "Recommended implementation details",
        "Problem summary",
        "Primary test type",
        "Component candidate",
        "CVT preconditions/mocks",
        "CVT steps",
        "CVT assertions",
        "E2E fallback",
        "Metadata gaps",
    ]
    table = [
        "",
        "## Closed bug-by-bug test blueprint",
        "",
        "|" + "|".join(table_headers) + "|",
        "|" + "|".join(["---"] * len(table_headers)) + "|",
    ]

    for row in rows:
        issue_cell = f"[#{row['issue_number']}]({row['issue_url']})"
        values = [
            issue_cell,
            row["created_at"],
            row["closed_at"],
            row["state_reason"],
            row["team_labels"],
            row["severity_labels"],
            row["real_bug_assessment"],
            row["real_bug_evidence"],
            row["ai_validity_verdict"],
            row["ai_validity_confidence"],
            row["ai_validity_analysis"],
            row["mobile_impact"],
            row["resolution_layer"],
            row["resolution_evidence"],
            row["canonical_issue"],
            row["recommended_test_type"],
            row["recommended_test_implementation"],
            row["problem_summary"],
            row["primary_test_type"],
            row["component_candidate"],
            row["cvt_preconditions_mocks"],
            row["cvt_steps"],
            row["cvt_assertions"],
            row["e2e_fallback"],
            row["metadata_gaps"],
        ]
        table.append("|" + "|".join(markdown_escape(str(v)) for v in values) + "|")

    MD_PATH.write_text("\n".join(header + table) + "\n")


def write_excel(rows: list[dict], generated_at: dt.datetime) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Closed Bug Matrix"

    headers = [
        "Issue #",
        "Issue URL",
        "Created (UTC)",
        "Closed (UTC)",
        "State reason",
        "Team labels",
        "Severity labels",
        "Real bug assessment",
        "Real bug evidence",
        "AI validity verdict",
        "AI confidence",
        "AI analysis",
        "Applies to mobile",
        "Resolution layer",
        "Resolution evidence",
        "Canonical issue for test",
        "Recommended test implementation",
        "Recommended implementation details",
        "Problem summary",
        "Primary test type",
        "Component candidate",
        "CVT preconditions/mocks",
        "CVT steps",
        "CVT assertions",
        "E2E fallback",
        "Metadata gaps",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(
            [
                row["issue_number"],
                row["issue_url"],
                row["created_at"],
                row["closed_at"],
                row["state_reason"],
                row["team_labels"],
                row["severity_labels"],
                row["real_bug_assessment"],
                row["real_bug_evidence"],
                row["ai_validity_verdict"],
                row["ai_validity_confidence"],
                row["ai_validity_analysis"],
                row["mobile_impact"],
                row["resolution_layer"],
                row["resolution_evidence"],
                row["canonical_issue"],
                row["recommended_test_type"],
                row["recommended_test_implementation"],
                row["problem_summary"],
                row["primary_test_type"],
                row["component_candidate"],
                row["cvt_preconditions_mocks"],
                row["cvt_steps"],
                row["cvt_assertions"],
                row["e2e_fallback"],
                row["metadata_gaps"],
            ]
        )

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 10,
        "B": 36,
        "C": 22,
        "D": 22,
        "E": 14,
        "F": 26,
        "G": 20,
        "H": 30,
        "I": 40,
        "J": 24,
        "K": 12,
        "L": 46,
        "M": 16,
        "N": 26,
        "O": 42,
        "P": 14,
        "Q": 34,
        "R": 56,
        "S": 50,
        "T": 36,
        "U": 38,
        "V": 54,
        "W": 44,
        "X": 44,
        "Y": 44,
        "Z": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    sev_counter = Counter(row["severity_top"] for row in rows)
    realness_counter = Counter(row["real_bug_assessment"] for row in rows)
    ai_verdict_counter = Counter(row["ai_validity_verdict"] for row in rows)
    resolution_counter = Counter(row["resolution_layer"] for row in rows)
    test_type_counter = Counter(row["recommended_test_type"] for row in rows)
    team_counter = Counter()
    for row in rows:
        for team in [t.strip() for t in row["team_labels"].split(",") if t.strip() and t.strip() != "Unlabeled"]:
            team_counter[team] += 1

    summary = wb.create_sheet("Summary")
    summary.append(["Generated at (UTC)", generated_at.isoformat(timespec="seconds")])
    summary.append(["Repository", REPO])
    summary.append(["Lookback days", LOOKBACK_DAYS])
    summary.append(["Filter", "state:CLOSED + label:type-bug"])
    summary.append(["Total bugs", len(rows)])
    summary.append([])

    summary.append(["Severity", "Count"])
    for severity, count in sev_counter.most_common():
        summary.append([severity, count])
    summary.append([])

    summary.append(["Real bug assessment", "Count"])
    for assessment, count in realness_counter.most_common():
        summary.append([assessment, count])
    summary.append([])

    summary.append(["AI validity verdict", "Count"])
    for verdict, count in ai_verdict_counter.most_common():
        summary.append([verdict, count])
    summary.append([])

    summary.append(["Resolution layer", "Count"])
    for layer, count in resolution_counter.most_common():
        summary.append([layer, count])
    summary.append([])

    summary.append(["Recommended test implementation", "Count"])
    for test_impl, count in test_type_counter.most_common():
        summary.append([test_impl, count])
    summary.append([])

    summary.append(["Team label", "Count"])
    for team, count in team_counter.most_common(20):
        summary.append([team, count])

    header_rows = []
    row_idx = 7
    header_rows.append(row_idx)
    row_idx += len(sev_counter) + 2
    header_rows.append(row_idx)
    row_idx += len(realness_counter) + 2
    header_rows.append(row_idx)
    row_idx += len(ai_verdict_counter) + 2
    header_rows.append(row_idx)
    row_idx += len(resolution_counter) + 2
    header_rows.append(row_idx)
    row_idx += len(test_type_counter) + 2
    header_rows.append(row_idx)
    for row_idx in header_rows:
        for cell in summary[row_idx]:
            cell.fill = header_fill
            cell.font = header_font

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 28
    wb.save(XLSX_PATH)


def main() -> None:
    generated_at = dt.datetime.now(dt.timezone.utc)
    issues = run_gh_issue_list()
    rows = build_rows(issues)
    write_markdown(rows, generated_at)
    write_excel(rows, generated_at)
    print(f"Generated rows: {len(rows)}")
    print(f"Markdown: {MD_PATH}")
    print(f"Excel: {XLSX_PATH}")


if __name__ == "__main__":
    main()
